import sys
import traceback
from PyQt6 import QtWidgets, QtCore, QtGui
from openpyxl.drawing.image import Image  # Добавляем импорт для работы с изображениями
from openpyxl.worksheet.page import PageMargins  # Для настройки полей
from ui.deals import Ui_MainWindow as DealForm
from datetime import datetime
from database import Database
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io


class DealWindow(QtWidgets.QMainWindow, DealForm):
    def __init__(self, user_email):
        super().__init__()
        self.setupUi(self)
        self.user_email = user_email
        self.current_user = None
        self.deal_data = None
        self.selected_deal_id = None
        self.db = Database()

        self.init_ui()
        self.load_user_data()
        self.load_deals()

        self.pushButton.clicked.connect(self.add_new_deal)
        self.pushButton_2.clicked.connect(self.logout)
        self.home.clicked.connect(self.go_home)
        self.task.clicked.connect(self.go_task)
        self.client.clicked.connect(self.go_client)
        self.empl.clicked.connect(self.go_empl)
        self.date.clicked.connect(self.go_date)
        self.sett.clicked.connect(self.go_sett)
        self.chat.clicked.connect(self.go_chat)
        self.ana.clicked.connect(self.go_ana)
        self.notification.clicked.connect(self.show_notifications)
        self.help.clicked.connect(self.show_help)

        self.tableWidget.cellDoubleClicked.connect(self.show_deal_details)

    def init_ui(self):
        """Настройка интерфейса"""
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels([
            "ID", "Название", "Тип", "Статус", "Организация", "Исполнитель"
        ])
        self.tableWidget.setColumnWidth(0, 50)
        self.tableWidget.setColumnWidth(1, 200)
        self.tableWidget.setColumnWidth(2, 150)
        self.tableWidget.setColumnWidth(3, 150)
        self.tableWidget.setColumnWidth(4, 150)
        self.tableWidget.setColumnWidth(5, 150)
        self.tableWidget.setColumnHidden(0, True)

    def load_user_data(self):
        """Загрузка данных пользователя"""
        try:
            self.current_user = self.db.get_user_by_email(self.user_email)
            if self.current_user:
                self.profil.setText(
                    f"{self.current_user[1]} {self.current_user[2]}\n"
                    f"{self.current_user[6]} ({self.current_user[5]})"
                )

                if self.current_user[5].lower() != 'менеджер':
                    self.pushButton.setVisible(False)
        except Exception as e:
            print(f"Ошибка загрузки данных пользователя: {e}")

    def load_deals(self):
        """Загрузка сделок в таблицу"""
        try:
            if self.current_user[5].lower() == 'менеджер':
                deals = self.db.get_deals_by_executor(self.current_user[0])
            else:
                deals = self.db.get_all_deals()

            self.deal_data = deals
            self.tableWidget.setRowCount(len(deals))

            for row, deal in enumerate(deals):
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(str(deal[0])))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(deal[1]))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(deal[2]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(deal[3]))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(deal[4]))
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(deal[5]))

                if deal[3] == "Не обработан":
                    color = QtGui.QColor(255, 200, 200)
                elif deal[3] == "Обработка":
                    color = QtGui.QColor(255, 255, 200)
                elif deal[3] == "Выставлен счёт/КП":
                    color = QtGui.QColor(200, 255, 200)
                elif deal[3] == "Оплата":
                    color = QtGui.QColor(200, 200, 255)
                elif deal[3] == "В производстве":
                    color = QtGui.QColor(200, 255, 255)
                elif deal[3] == "Завершен":
                    color = QtGui.QColor(200, 200, 200)

                for col in range(6):
                    self.tableWidget.item(row, col).setBackground(color)
        except Exception as e:
            print(f"Ошибка загрузки сделок: {e}")

    def show_deal_details(self, row, column):
        """Показ деталей сделки"""
        try:
            self.selected_deal_id = int(self.tableWidget.item(row, 0).text())
            deal = next((d for d in self.deal_data if d[0] == self.selected_deal_id), None)
            if not deal:
                return

            # Проверяем, что в deal достаточно элементов
            if len(deal) < 12:  # Если в кортеже меньше 12 элементов
                # Создаем новый кортеж с недостающими значениями, заполненными None
                deal = deal + (None,) * (12 - len(deal))

            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle("Детали сделки")
            dialog.setMinimumSize(500, 400)

            layout = QtWidgets.QVBoxLayout(dialog)
            form_layout = QtWidgets.QFormLayout()

            name_label = QtWidgets.QLabel(deal[1])
            name_label.setStyleSheet("font-weight: bold; font-size: 16px;")
            form_layout.addRow("Название:", name_label)
            form_layout.addRow("Тип сделки:", QtWidgets.QLabel(deal[2]))

            status_label = QtWidgets.QLabel(deal[3])
            form_layout.addRow("Статус:", status_label)
            form_layout.addRow("Организация:", QtWidgets.QLabel(deal[4]))
            form_layout.addRow("Исполнитель:", QtWidgets.QLabel(deal[5]))

            date1 = deal[6].strftime("%d.%m.%Y %H:%M") if deal[6] else "Не указана"
            date2 = deal[7].strftime("%d.%m.%Y %H:%M") if deal[7] else "Не указана"
            form_layout.addRow("Дата создания:", QtWidgets.QLabel(date1))
            form_layout.addRow("Дата завершения:", QtWidgets.QLabel(date2))

            # Показываем финансовую информацию для бухгалтеров, админов или менеджеров (если их сделка)
            show_finance = (
                    self.current_user[5].lower() == 'бухгалтер' or
                    self.current_user[5].lower() == 'админ' or
                    (self.current_user[5].lower() == 'менеджер' and
                     deal[5] == f"{self.current_user[1]} {self.current_user[2]} {self.current_user[3]}")
            )

            if show_finance:
                price = f"{deal[8]:.2f} руб." if deal[8] is not None else "Не указана"
                nds = f"{deal[9]}%" if deal[9] is not None else "Не указан"
                total = f"{deal[10]:.2f} руб." if deal[10] is not None else "Не указана"

                form_layout.addRow("Цена:", QtWidgets.QLabel(price))
                form_layout.addRow("НДС:", QtWidgets.QLabel(nds))
                form_layout.addRow("Итоговая сумма:", QtWidgets.QLabel(total))

            # Добавляем возможность просмотра счета для тех, кто может видеть финансовую информацию
            if show_finance and deal[11] is not None:  # Проверяем наличие счета (поле 11)
                view_bill_btn = QtWidgets.QPushButton("Просмотреть счет")
                view_bill_btn.clicked.connect(lambda: self.view_bill(deal[11]))
                form_layout.addRow("Счет:", view_bill_btn)

            layout.addLayout(form_layout)
            button_box = QtWidgets.QDialogButtonBox()

            # Бухгалтер может загрузить счет и указать цену только в определенных статусах
            if (self.current_user[5].lower() == 'бухгалтер' and
                    deal[3] in ["Обработка", "Выставлен счёт/КП"]):

                # Проверяем, можно ли загружать счет (если еще не загружен)
                if deal[11] is None:
                    upload_bill_btn = button_box.addButton("Загрузить счет",
                                                           QtWidgets.QDialogButtonBox.ButtonRole.ActionRole)
                    upload_bill_btn.clicked.connect(lambda: self.upload_bill(dialog))

                # Проверяем, можно ли указывать цену (только в статусе "Обработка" и если цена не указана)
                if deal[3] == "Обработка" and deal[8] is None:
                    set_price_btn = button_box.addButton("Указать цену",
                                                         QtWidgets.QDialogButtonBox.ButtonRole.ActionRole)
                    set_price_btn.clicked.connect(lambda: self.set_price_nds(dialog))

            # Менеджер может менять статус только своих сделок
            elif (self.current_user[5].lower() == 'менеджер' and
                  deal[5] == f"{self.current_user[1]} {self.current_user[2]} {self.current_user[3]}"):
                status_combo = QtWidgets.QComboBox()
                statuses = self.db.get_all_deal_statuses()
                for status_id, status_name in statuses:
                    status_combo.addItem(status_name, status_id)

                current_index = status_combo.findText(deal[3])
                if current_index >= 0:
                    status_combo.setCurrentIndex(current_index)

                change_status_btn = button_box.addButton("Изменить статус",
                                                         QtWidgets.QDialogButtonBox.ButtonRole.ActionRole)
                change_status_btn.clicked.connect(lambda: self.change_deal_status(status_combo.currentData(), dialog))

                form_layout.addRow("Новый статус:", status_combo)

            close_btn = button_box.addButton("Закрыть", QtWidgets.QDialogButtonBox.ButtonRole.RejectRole)
            close_btn.clicked.connect(dialog.reject)

            layout.addWidget(button_box)
            dialog.exec()
        except Exception as e:
            print(f"Ошибка показа деталей сделки: {e}")
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить детали сделки: {str(e)}")

    def generate_bill(self, deal_id):
        """Генерация счета с корректным отображением чисел и изображениями"""
        try:
            # 1. Получаем данные сделки
            deal = next((d for d in self.deal_data if d[0] == deal_id), None)
            if not deal:
                QtWidgets.QMessageBox.warning(self, "Ошибка", "Сделка не найдена")
                return None

            # 2. Получаем данные организации
            org_data = self.db.get_organization_details(deal[4])
            if not org_data:
                QtWidgets.QMessageBox.warning(self, "Ошибка", "Организация не найдена")
                return None

            # 3. Подготавливаем данные
            supplier = {
                'name': "ООО «НГ-СОФТ»",
                'inn': "9729293217",
                'kpp': "771701001",
                'address': "г. Москва, ул. 1-я Мытищинская, д. 28, стр. 1",
                'phone': "+7 (495) 682-26-20",
                'director': "Сукиасян Р.Х."
            }

            client = {
                'name': org_data.get('name', "Без названия"),
                'inn': org_data.get('inn', "не указан"),
                'kpp': org_data.get('kpp', "не указан"),
                'address': org_data.get('address', "не указан")
            }

            # 4. Финансовые расчеты
            price = float(deal[8]) if deal[8] else 0.0
            nds = int(deal[9]) if deal[9] else 20
            nds_sum = round(price * nds / 100, 2)
            total = round(price + nds_sum, 2)

            # 5. Создаем Excel-документ
            wb = Workbook()
            ws = wb.active
            ws.title = "Счет"

            # 6. Настройка стилей
            bold = Font(bold=True, name='Arial', size=10)
            header_font = Font(bold=True, name='Arial', size=12)
            regular = Font(name='Arial', size=10)
            center = Alignment(horizontal='center', vertical='center')
            right = Alignment(horizontal='right', vertical='center')
            left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            border_thick = Border(bottom=Side(style='thick'))

            # 7. Устанавливаем ширину столбцов
            column_widths = {
                'A': 5,  # №
                'B': 40,  # Наименование
                'C': 8,  # Кол-во
                'D': 5,  # Ед.
                'E': 15,  # Цена
                'F': 8,  # НДС
                'G': 15,  # Сумма НДС
                'H': 18  # Всего (увеличено для больших сумм)
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 8. Заполнение шапки
            # Номер счета и дата
            ws.merge_cells('A1:H1')
            ws['A1'] = f"Счет № {deal_id} от {datetime.now().strftime('%d.%m.%Y')}"
            ws['A1'].font = header_font
            ws['A1'].alignment = center

            # Поставщик
            ws.merge_cells('A3:B3')
            ws['A3'] = "Поставщик:"
            ws['A3'].font = bold

            ws.merge_cells('C3:H3')
            ws['C3'] = (f"{supplier['name']}\n"
                        f"ИНН {supplier['inn']} КПП {supplier['kpp']}\n"
                        f"Адрес: {supplier['address']}\n"
                        f"Тел.: {supplier['phone']}")
            ws['C3'].font = regular
            ws['C3'].alignment = left
            ws.row_dimensions[3].height = 60

            # Покупатель
            ws.merge_cells('A5:B5')
            ws['A5'] = "Покупатель:"
            ws['A5'].font = bold

            ws.merge_cells('C5:H5')
            ws['C5'] = (f"{client['name']}\n"
                        f"ИНН {client['inn']} КПП {client['kpp']}\n"
                        f"Адрес: {client['address']}")
            ws['C5'].font = regular
            ws['C5'].alignment = left
            ws.row_dimensions[5].height = 45

            # 9. Таблица товаров
            headers = ["№", "Товары (работы, услуги)", "Кол-во", "Ед.", "Цена", "НДС", "Сумма НДС", "Всего"]

            # Заголовки таблицы
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = bold
                cell.border = border
                cell.alignment = center

            # Данные товара
            data_row = [
                1,  # №
                deal[1],  # Наименование
                1,  # Кол-во
                "шт.",  # Ед.
                price,  # Цена
                f"{nds}%",  # НДС
                nds_sum,  # Сумма НДС
                total  # Всего
            ]

            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=8, column=col, value=value)
                cell.font = regular
                cell.border = border
                cell.alignment = center if col not in [2, 5, 7, 8] else right

                # Форматирование числовых колонок
                if col in [5, 7, 8]:
                    cell.number_format = '#,##0.00'
                    # Автоматическое расширение столбца для больших чисел
                    if len(f"{value:,.2f}") > ws.column_dimensions[chr(64 + col)].width:
                        ws.column_dimensions[chr(64 + col)].width = len(f"{value:,.2f}") + 2

            # 10. Итого и Всего к оплате
            ws.merge_cells('A10:G10')
            ws['A10'] = "Итого:"
            ws['A10'].font = bold
            ws['H10'] = total
            ws['H10'].font = bold
            ws['H10'].number_format = '#,##0.00'
            ws['H10'].border = border_thick

            ws.merge_cells('A12:G12')
            ws['A12'] = "Всего к оплате:"
            ws['A12'].font = header_font
            ws['H12'] = total
            ws['H12'].font = header_font
            ws['H12'].number_format = '#,##0.00'
            ws['H12'].border = border_thick

            # 11. Подпись и печать
            try:
                # Печать
                pech_img = Image("res/pech.png")
                pech_img.width = 100
                pech_img.height = 100
                ws.add_image(pech_img, "D15")

                # Подпись
                pod_img = Image("res/pod.png")
                pod_img.width = 100
                pod_img.height = 50
                ws.add_image(pod_img, "F15")
            except Exception as img_error:
                print(f"Ошибка загрузки изображений: {img_error}")
                # Запасной вариант
                ws['D15'] = "[МЕСТО ДЛЯ ПЕЧАТИ]"
                ws['F15'] = "[МЕСТО ДЛЯ ПОДПИСИ]"
                ws['D15'].font = regular
                ws['F15'].font = regular

            # Подписи
            ws['D17'] = "М.П."
            ws['D17'].alignment = center
            ws['F17'] = supplier['director']
            ws['F17'].alignment = center

            # 12. Настройка страницы
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4

            # 13. Сохранение
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Ошибка",
                                           f"Ошибка генерации счета:\n{str(e)}")
            return None

    def upload_bill(self, dialog):
        """Генерация и сохранение счета"""
        try:
            # Получаем данные сделки
            deal = next((d for d in self.deal_data if d[0] == self.selected_deal_id), None)

            # Проверяем наличие названия организации
            if not deal[4] or not deal[4].strip():
                QtWidgets.QMessageBox.warning(self, "Ошибка",
                                              "Не указано название организации в сделке. Сначала заполните данные организации.")
                return

            # Проверяем существование организации в базе
            org_data = self.db.get_organization_details(deal[4])
            if not org_data:
                QtWidgets.QMessageBox.warning(self, "Ошибка",
                                              f"Организация '{deal[4]}' не найдена в базе данных.\n"
                                              "Проверьте правильность названия или добавьте организацию в справочник.")
                return

            # Проверяем наличие цены
            if deal[8] is None:
                QtWidgets.QMessageBox.warning(self, "Ошибка",
                                              "Не указана цена в сделке. Сначала укажите цену.")
                return

            # Генерируем счет
            bill_data = self.generate_bill(self.selected_deal_id)
            if not bill_data:
                return

            # Сохраняем в базу данных
            if self.db.update_deal_bill(self.selected_deal_id, bill_data):
                QtWidgets.QMessageBox.information(self, "Успех",
                                                  "Счет успешно сгенерирован и сохранен")
                dialog.accept()
                self.load_deals()

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка",
                                           f"Произошла ошибка при генерации счета: {str(e)}")

    def view_bill(self, bill_data):
        """Просмотр счета"""
        try:
            # Создаем временный файл для просмотра
            import tempfile
            import os
            import subprocess

            # Определяем расширение файла
            file_extension = '.xlsx'  # Так как мы генерируем только Excel-файлы

            with tempfile.NamedTemporaryFile(suffix=file_extension, delete=False) as temp_file:
                temp_file.write(bill_data)
                temp_file_path = temp_file.name

            # Открываем файл с помощью стандартного приложения системы
            if sys.platform == "win32":
                os.startfile(temp_file_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", temp_file_path])
            else:
                subprocess.run(["xdg-open", temp_file_path])

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось открыть счет: {str(e)}")

    def set_price_nds(self, dialog):
        """Установка цены и НДС"""
        try:
            price, ok = QtWidgets.QInputDialog.getDouble(
                self, "Укажите цену", "Цена (руб):", 0, 0, 9999999, 2
            )
            if ok:
                nds, ok = QtWidgets.QInputDialog.getInt(
                    self, "Укажите НДС", "НДС (%):", 20, 0, 100
                )
                if ok:
                    total_price = price * (1 + nds / 100)
                    if self.db.update_deal_price(
                            self.selected_deal_id, price, nds, total_price, "Выставлен счёт/КП"
                    ):
                        QtWidgets.QMessageBox.information(self, "Успех", "Данные успешно обновлены")
                        dialog.accept()
                        self.load_deals()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def change_deal_status(self, new_status_id, dialog):
        """Изменение статуса сделки"""
        try:
            if self.db.update_deal_status(self.selected_deal_id, new_status_id):
                QtWidgets.QMessageBox.information(self, "Успех", "Статус изменен")
                dialog.accept()
                self.load_deals()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def add_new_deal(self):
        """Добавление новой сделки"""
        try:
            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle("Новая сделка")
            dialog.setMinimumSize(400, 300)

            layout = QtWidgets.QVBoxLayout(dialog)
            form_layout = QtWidgets.QFormLayout()

            name_edit = QtWidgets.QLineEdit()
            form_layout.addRow("Название:", name_edit)

            type_combo = QtWidgets.QComboBox()
            for type_data in self.db.get_all_deal_types():
                type_combo.addItem(type_data[1], type_data[0])  # type_data[1] - name, type_data[0] - id
            form_layout.addRow("Тип сделки:", type_combo)

            org_combo = QtWidgets.QComboBox()
            for org_data in self.db.get_all_organizations():
                org_combo.addItem(org_data[1], org_data[0])  # org_data[1] - name, org_data[0] - id
            form_layout.addRow("Организация:", org_combo)

            date_edit = QtWidgets.QDateEdit()
            date_edit.setCalendarPopup(True)
            date_edit.setDate(QtCore.QDate.currentDate().addMonths(1))
            form_layout.addRow("Дата завершения:", date_edit)

            layout.addLayout(form_layout)
            button_box = QtWidgets.QDialogButtonBox()

            save_btn = button_box.addButton("Сохранить", QtWidgets.QDialogButtonBox.ButtonRole.AcceptRole)
            save_btn.clicked.connect(lambda: self.save_new_deal(
                name_edit.text(),
                type_combo.currentData(),
                org_combo.currentData(),
                date_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss"),
                dialog
            ))

            cancel_btn = button_box.addButton("Отмена", QtWidgets.QDialogButtonBox.ButtonRole.RejectRole)
            cancel_btn.clicked.connect(dialog.reject)

            layout.addWidget(button_box)
            dialog.exec()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
            print(e)
    def save_new_deal(self, name, type_id, org_id, end_date, dialog):
        """Сохранение новой сделки"""
        try:
            if not name:
                QtWidgets.QMessageBox.warning(self, "Ошибка", "Укажите название")
                return

            if self.db.create_deal(
                    name=name,
                    type_id=type_id,
                    org_id=org_id,
                    executor_id=self.current_user[0],
                    end_date=end_date
            ):
                QtWidgets.QMessageBox.information(self, "Успех", "Сделка создана")
                dialog.accept()
                self.load_deals()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
            print(e)


    def logout(self):
        """Выход из системы"""
        from LoginApp import Authorization
        self.auth_window = Authorization()
        self.auth_window.show()
        self.close()

    def go_home(self):
        """Переход на главный экран"""
        from HomeApp import HomeWindow
        self.home_window = HomeWindow(self.user_email)
        self.home_window.show()
        self.close()

    def go_task(self):
        """Переход к задачам"""
        from TaskApp import TaskWindow
        self.task_window = TaskWindow(self.user_email)
        self.task_window.show()
        self.close()

    def go_client(self):
        """Переход к клиентам"""
        from ClientApp import ClientWindow
        self.client_window = ClientWindow(self.user_email)
        self.client_window.show()
        self.close()

    def go_empl(self):
        """Переход к сотрудникам"""
        from EmplApp import EmplWindow
        self.empl_window = EmplWindow(self.user_email)
        self.empl_window.show()
        self.close()

    def go_date(self):
        # Выход из системы
        from DateApp import DateWindow
        self.date_window = DateWindow(self.user_email)
        self.date_window.show()
        self.close()

    def go_sett(self):
        from SetApp import SetWindow
        self.set_window = SetWindow(self.user_email)
        self.set_window.show()
        self.close()

    def go_chat(self):
        # Выход из системы
        from ChatApp import ChatWindow
        self.chat_window = ChatWindow(self.user_email)
        self.chat_window.show()
        self.close()

    def go_ana(self):
        from AnApp import AnWindow
        self.ana_window = AnWindow(self.user_email)
        self.ana_window.show()
        self.close()

    def show_notifications(self):
        from NotificationApp import NotificationWindow
        """Показать уведомления"""
        self.notification_window = NotificationWindow(self.user_email)
        self.notification_window.exec()

    def show_help(self):
        from HelpApp import HelpWindow
        """Показать справку"""
        self.help_window = HelpWindow()
        self.help_window.exec()

    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        self.db.close()
        event.accept()


